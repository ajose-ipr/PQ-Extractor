import streamlit as st
import os
import re
import pdfplumber
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Set page config
st.set_page_config(
    page_title="PQ Report Harmonics Extractor",
    page_icon="📋",
    layout="wide"
)

st.title("📋 PQ Report Harmonics Extractor")

# Constants
EXPECTED_HARMONICS = set(range(2, 51))
FAIL_COLOR = 'color: darkred; font-weight: bold; background-color: #ffebeb'
HIGHLIGHT_COLOR = 'font-weight: bold; background-color: #fff3cd'

# Column definitions for all table types
VOLTAGE_COLUMNS = [
    "N", "[%]", "Reg Max[%]",
    "Measured_V1N", "Measured_V2N", "Measured_V3N", 
    "Result_V1N", "Result_V2N", "Result_V3N"
]
CURRENT_COLUMNS = [
    "N", "[%]", "Reg Max[%]",
    "Measured_I1", "Measured_I2", "Measured_I3", 
    "Result_I1", "Result_I2", "Result_I3"
]

# Updated section boundaries for all tables - FIXED TO EXCLUDE HARMONIC 5:
SECTION_BOUNDARIES = {
    "HARMONIC VOLTAGE FULL TIME RANGE": [
        "SUMMARY", "TOTAL HARMONIC VOLTAGE FULL TIME RANGE", 
        "TOTAL HARMONIC DISTORTION FULL TIME RANGE", "HARMONIC CURRENT FULL TIME RANGE"
    ],
    "HARMONIC CURRENT FULL TIME RANGE": [
        "TOTAL HARMONIC DISTORTION DAILY", "TDD FULL TIME RANGE",
        "HARMONIC VOLTAGE DAILY", "TRANSIENT"
    ],
    "HARMONIC VOLTAGE DAILY": [
        "TOTAL HARMONIC DISTORTION FULL TIME RANGE", 
        "TOTAL HARMONIC VOLTAGE FULL TIME RANGE", "HARMONIC CURRENT DAILY", "TOTAL HARMONIC DISTORTION DAILY"
    ],
    "HARMONIC CURRENT DAILY": [
        "TDD FULL TIME RANGE", "TDD DAILY", "TRANSIENT", "FLICKER SEVERITY"
    ]
}

# All supported table names
SUPPORTED_TABLES = [
    "Harmonic Voltage Full Time Range",
    "Harmonic Current Full Time Range", 
    "Harmonic Voltage Daily",
    "Harmonic Current Daily"
]

# Enhanced regex patterns for text extraction
TEXT_EXTRACTION_PATTERNS = [
    # Standard pattern with Pass/Fail
    re.compile(
        r'(\d+)\s*,?\s*'  # Harmonic
        r'(\d+)\s*,?\s*'  # Time percent (%)
        r'([\d.]+)\s*,?\s*'  # Reg max
        r'([\d.]+)\s*,?\s*'  # Measured 1
        r'([\d.]+)\s*,?\s*'  # Measured 2
        r'([\d.]+)\s*,?\s*'  # Measured 3
        r'(Pass|Fail)\s*\(([\d.%]+)\)\s*,?\s*'  # Result 1
        r'(Pass|Fail)\s*\(([\d.%]+)\)\s*,?\s*'  # Result 2
        r'(Pass|Fail)\s*\(([\d.%]+)\)',  # Result 3
        re.IGNORECASE
    ),
    # Pattern without explicit Pass/Fail words
    re.compile(
        r'(\d+)\s*,?\s*'
        r'(\d+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'([\d.]+)\s*,?\s*'
        r'\(([\d.%]+)\)\s*,?\s*'
        r'\(([\d.%]+)\)\s*,?\s*'
        r'\(([\d.%]+)\)',
        re.IGNORECASE
    ),
    # Pattern for multiline data
    re.compile(
        r'(\d+)\s*,?\s*(\d+)\s*,?\s*([\d.]+)\s*,?\s*([\d.]+)\s*,?\s*([\d.]+)\s*,?\s*([\d.]+)',
        re.IGNORECASE | re.MULTILINE
    )
]

# UI Setup
st.sidebar.header("📂 Select Input Method")
uploaded_files = st.sidebar.file_uploader(
    "Upload all PDF files from your week report folder", 
    type="pdf", 
    accept_multiple_files=True,
    help="Select all 15 files: 7 Days summary + Day 1-7 Day/Night reports"
)

selected_file = None
if uploaded_files:
    st.sidebar.subheader("📋 Uploaded PDF Files")
    for uploaded_file in uploaded_files:
        if st.sidebar.button(uploaded_file.name, key=f"select_{uploaded_file.name}"):
            selected_file = uploaded_file
            st.session_state.selected_file = uploaded_file

if 'selected_file' in st.session_state:
    selected_file = st.session_state.selected_file

def extract_metadata(pdf_file, filename):
    """Extract metadata from PDF file"""
    name = filename if isinstance(filename, str) else filename.name
    component_info = re.findall(r"\((.*?)\)", str(name))
    component_text = component_info[0] if component_info else "Not found"

    report_info = {
        "start_time": "Not found", 
        "end_time": "Not found", 
        "gmt": "Not found", 
        "version": "Not found"
    }

    try:
        with pdfplumber.open(pdf_file if isinstance(pdf_file, str) else pdf_file) as pdf:
            text0 = pdf.pages[0].extract_text() or ""
        
        time_pattern = re.compile(
            r"Start time:\s*(\d{2}-\d{2}-\d{4}\s*\d{2}:\d{2}:\d{2}\s*[AP]M)\s*"
            r"End time:\s*(\d{2}-\d{2}-\d{4}\s*\d{2}:\d{2}:\d{2}\s*[AP]M)\s*"
            r"GMT:\s*([+-]\d{2}:\d{2})\s*"
            r"Report Version:\s*([\d.]+)"
        )
        
        match = time_pattern.search(text0)
        if match:
            report_info = {
                "start_time": match.group(1),
                "end_time": match.group(2),
                "gmt": match.group(3),
                "version": match.group(4)
            }
            
        combined_text = (str(name) + " " + text0).upper()
        block_match = re.search(r"\bBLOCK[-\s]*(\d{1,3})\b", combined_text)
        feeder_match = re.search(r"\b(FEEDER|BAY)[-\s]*(\d{1,3})\b", combined_text)
        company_match = re.search(r"\b(TATA|ADANI|NTPC|RELIANCE|POWERGRID|TORRENT)\b", combined_text)

        block = block_match.group(1) if block_match else "Not found"
        feeder = feeder_match.group(2) if feeder_match else "Not found"
        company = company_match.group(1) if company_match else "Not found"

        return component_text, block, feeder, company, report_info

    except Exception as e:
        st.error(f"Error extracting metadata: {str(e)}")
        return "Not found", "Error", "Error", "Error", report_info

def extract_table_data_from_text(text, has_results=True):
    """Enhanced text extraction for all table types"""
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'(Pass|Fail)\s*\(\s*([\d.%]+)\s*\)', r'\1(\2)', text)
    
    data = []
    
    # Try each pattern
    for pattern in TEXT_EXTRACTION_PATTERNS:
        for match in pattern.finditer(text):
            try:
                harmonic = int(match.group(1))
                
                # FILTER 1: Skip harmonic 1 (fundamental frequency)
                if harmonic == 1:
                    continue
                
                # FILTER 2: Skip non-harmonic data (years, dates, etc.)
                # Valid harmonics are 2-50, anything outside this range is likely date/year data
                if harmonic < 2 or harmonic > 50:
                    continue
                
                # FILTER 3: Additional validation - if the "harmonic" looks like a year (>1000)
                if harmonic > 1000:
                    continue
                
                groups = match.groups()
                
                if len(groups) >= 12:  # Full pattern with Pass/Fail
                    row = [
                        harmonic, match.group(2), match.group(3), match.group(4),
                        match.group(5), match.group(6), f"{match.group(7)}({match.group(8)})",
                        f"{match.group(9)}({match.group(10)})", f"{match.group(11)}({match.group(12)})"
                    ]
                elif len(groups) >= 9:  # Pattern without Pass/Fail
                    row = [
                        harmonic, match.group(2), match.group(3), match.group(4),
                        match.group(5), match.group(6), f"Pass({match.group(7)})",
                        f"Pass({match.group(8)})", f"Pass({match.group(9)})"
                    ]
                elif len(groups) >= 6 and not has_results:  # Just measurements
                    row = [
                        harmonic, match.group(2), match.group(3), match.group(4),
                        match.group(5), match.group(6), "N/A", "N/A", "N/A"
                    ]
                else:
                    continue
                    
                data.append(row)
            except (ValueError, IndexError):
                continue
    
    return data

def extract_tables_from_pdf(file):
    """Extract all harmonic tables from PDF starting from page 2"""
    tables = {table_name: [] for table_name in SUPPORTED_TABLES}
    
    try:
        with pdfplumber.open(file if isinstance(file, str) else file) as pdf:
            active_table = None
            
            # Skip first page as requested
            for page_num, page in enumerate(pdf.pages):
                if page_num == 0:  # Skip first page
                    continue
                    
                page_text = page.extract_text() or ""
                upper_text = page_text.upper()
                page_tables = page.extract_tables()
                
                # Check for table headers
                for table_name in tables:
                    table_name_upper = table_name.upper()
                    if table_name_upper in upper_text:
                        start_idx = upper_text.find(table_name_upper)
                        end_idx = len(page_text)
                        
                        # Find section boundaries
                        for boundary in SECTION_BOUNDARIES.get(table_name_upper, []):
                            boundary_idx = upper_text.find(boundary, start_idx + len(table_name))
                            if boundary_idx != -1:
                                end_idx = min(end_idx, boundary_idx)
                        
                        section_text = page_text[start_idx:end_idx]
                        active_table = table_name
                        
                        # Extract structured tables
                        _extract_structured_data(page_tables, tables, active_table)
                        
                        # Extract from text as fallback
                        _extract_text_data(section_text, tables, active_table)
                        continue
                
                # Continue extracting for active table (FIXED BOUNDARY CHECK)
                if active_table and not _check_boundary_hit(upper_text, active_table):
                    _extract_structured_data(page_tables, tables, active_table)
                    _extract_text_data(page_text, tables, active_table)
                else:
                    # Only reset active_table if we actually hit a real boundary, not "HARMONIC 5:"
                    if active_table and _check_boundary_hit(upper_text, active_table):
                        # Special case: Don't stop for "HARMONIC 5:" when processing Harmonic Current Daily
                        if active_table == "Harmonic Current Daily" and "HARMONIC 5:" in upper_text:
                            # Continue processing this page for the current table
                            _extract_structured_data(page_tables, tables, active_table)
                            _extract_text_data(page_text, tables, active_table)
                        else:
                            active_table = None

    except Exception as e:
        st.error(f"Error processing PDF: {str(e)}")
    
    return tables

def _extract_structured_data(page_tables, tables, active_table):
    """Helper function to extract structured table data"""
    for table in page_tables:
        if len(table) > 1:
            for row in table:
                if row and str(row[0]).strip().isdigit():
                    try:
                        harmonic = int(row[0])
                        
                        # FILTER 1: Skip harmonic 1 (fundamental frequency)
                        if harmonic == 1:
                            continue
                            
                        # FILTER 2: Only accept valid harmonic range (2-50)
                        if harmonic < 2 or harmonic > 50:
                            continue
                            
                        # FILTER 3: Skip year-like numbers (>1000)
                        if harmonic > 1000:
                            continue
                        
                        if len(row) >= 9:
                            clean_row = [str(cell).strip() if cell is not None else "" for cell in row[:9]]
                            tables[active_table].append(clean_row)
                    except (ValueError, IndexError):
                        continue

def _extract_text_data(text, tables, active_table):
    """Helper function to extract text-based data"""
    text_data = extract_table_data_from_text(text)
    if text_data:
        existing_harmonics = {int(row[0]) for row in tables[active_table] if row and str(row[0]).isdigit()}
        for new_row in text_data:
            try:
                harmonic_value = int(new_row[0])
                
                # ADDITIONAL FILTER: Ensure we only add valid harmonics (2-50)
                if 2 <= harmonic_value <= 50 and harmonic_value not in existing_harmonics:
                    tables[active_table].append(new_row)
            except (ValueError, IndexError):
                continue

def _check_boundary_hit(upper_text, active_table):
    """Helper function to check if section boundary is hit - FIXED"""
    active_upper = active_table.upper()
    boundaries = SECTION_BOUNDARIES.get(active_upper, [])
    
    # For Harmonic Current Daily, exclude "HARMONIC 5:" from boundaries
    if active_upper == "HARMONIC CURRENT DAILY":
        boundaries = [b for b in boundaries if "HARMONIC 5" not in b.upper()]
    
    return any(boundary in upper_text for boundary in boundaries)

def process_table_data(table_data, table_name=None):
    """Process and validate table data"""
    columns = CURRENT_COLUMNS if table_name and "Current" in table_name else VOLTAGE_COLUMNS
    
    if not table_data:
        return pd.DataFrame(columns=columns)

    try:
        df = pd.DataFrame(table_data, columns=columns)
        df['N'] = pd.to_numeric(df['N'], errors='coerce')
        df = df.dropna(subset=['N'])
        
        # CRITICAL FILTER: Remove fundamental frequency and invalid harmonics
        df = df[df['N'] != 1]
        
        # ADDITIONAL FILTER: Only keep valid harmonic range (2-50)
        # This removes any year data (2024, 2025, etc.) that might have been captured
        df = df[(df['N'] >= 2) & (df['N'] <= 50)]
        
        df = df.drop_duplicates(subset=['N', '[%]'])
        
        found_harmonics = set(df['N'].astype(int))
        missing = sorted(EXPECTED_HARMONICS - found_harmonics)
        
        # for global harmonics check
        #if missing:
        #    st.warning(f"Missing harmonics in {table_name}: {missing[:10]}{'...' if len(missing) > 10 else ''}")
        
        numeric_cols = ["N", "[%]", "Reg Max[%]", columns[3], columns[4], columns[5]]
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        
        return df.dropna()
    
    except Exception as e:
        st.error(f"Error processing table data: {str(e)}")
        return pd.DataFrame(columns=columns)

def split_table(df):
    """Split table by time limits and odd/even harmonics"""
    if df.empty:
        return {"95": (pd.DataFrame(), pd.DataFrame()), "99": (pd.DataFrame(), pd.DataFrame())}
    
    def split_odd_even(df_subset):
        if df_subset.empty:
            return pd.DataFrame(), pd.DataFrame()
        df_subset = df_subset.sort_values("N")
        odd = df_subset[df_subset["N"] % 2 == 1].reset_index(drop=True)
        even = df_subset[df_subset["N"] % 2 == 0].reset_index(drop=True)
        return odd, even
    
    df_95 = df[df["[%]"] == 95.0].copy()
    df_99 = df[df["[%]"] == 99.0].copy()
    
    return {"95": split_odd_even(df_95), "99": split_odd_even(df_99)}

def highlight_fails(df):
    """Apply styling to highlight failed measurements"""
    if df.empty:
        return df

    reg_max_col = "Reg Max[%]"
    measured_cols = [col for col in df.columns if col.startswith("Measured_")]
    result_cols = [col for col in df.columns if col.startswith("Result_")]
    
    def extract_value(val):
        """Extract numeric value from various string formats"""
        try:
            if isinstance(val, str):
                num_str = re.search(r'(\d+\.?\d*)', val.replace('%', '')).group(1)
                return float(num_str)
            return float(val)
        except:
            return 0.0

    def apply_row(row):
        styles = {col: '' for col in df.columns}
        
        try:
            threshold = float(row[reg_max_col]) if pd.notna(row.get(reg_max_col)) else 0.0
        except:
            threshold = 0.0
            
        # Check measurements and results
        for i, col in enumerate(measured_cols):
            try:
                value = extract_value(row[col])
                result_col = result_cols[i] if i < len(result_cols) else None
                
                # Check if fail in result or value exceeds threshold
                fail_condition = value > threshold
                if result_col and pd.notna(row.get(result_col)):
                    result_str = str(row[result_col]).lower()
                    if 'fail' in result_str:
                        fail_condition = True
                
                if fail_condition:
                    styles[col] = FAIL_COLOR
                    if result_col:
                        styles[result_col] = FAIL_COLOR
                    styles['N'] = HIGHLIGHT_COLOR
            except:
                continue
                
        return pd.Series(styles)

    return df.style.apply(apply_row, axis=1).set_properties(
        **{'text-align': 'center', 'border': '1px solid #ddd'}
    ).format("{:g}", subset=measured_cols)

def display_table_section(title, odd_df, even_df):
    """Display tables in two columns for odd/even harmonics with fail highlighting"""
    if title.strip():  # Only show title if it's not empty or just spaces
        st.markdown(f"### {title}")
    col1, col2 = st.columns(2)

    def display_harmonics_table(df, harmonic_type, column):
        with column:
            st.markdown(f"<h4 style='color: grey;'>{harmonic_type} Harmonics </h4>", unsafe_allow_html=True)
            if not df.empty:
                styled_df = highlight_fails(df)
                st.dataframe(styled_df, height=400, use_container_width=True)

                found = set(df['N'].astype(int))
                expected = set(h for h in EXPECTED_HARMONICS if h % 2 == (1 if harmonic_type == "Odd" else 0))
                missing = sorted(expected - found)
                if missing:
                    st.info(f"Missing {harmonic_type.lower()} harmonics: {missing[:10]}{'...' if len(missing) > 10 else ''}")
            else:
                st.warning(f"No {harmonic_type.lower()} harmonics data available")

    display_harmonics_table(odd_df, "Odd", col1)
    display_harmonics_table(even_df, "Even", col2)

def highlight_fails_in_excel(df, ws, start_row=1):
    """Apply conditional formatting to highlight failed harmonics in Excel"""
    if df.empty:
        return
    
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fail_font = Font(color='9C0006', bold=True)
    harmonic_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    reg_max_col = None
    measured_cols = []
    result_cols = []
    
    for idx, col in enumerate(df.columns, 1):
        if "Reg Max" in col:
            reg_max_col = idx
        elif col.startswith(('Measured_')):
            measured_cols.append(idx)
        elif col.startswith('Result_'):
            result_cols.append(idx)
    
    if not reg_max_col or not measured_cols:
        return
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        if r_idx == start_row:
            continue
            
        try:
            threshold = float(ws.cell(row=r_idx, column=reg_max_col).value)
        except (ValueError, TypeError):
            threshold = 0.0
        
        for i, m_col in enumerate(measured_cols):
            try:
                cell = ws.cell(row=r_idx, column=m_col)
                value = float(cell.value) if cell.value else 0.0
                
                result_value = None
                if i < len(result_cols):
                    result_cell = ws.cell(row=r_idx, column=result_cols[i])
                    result_value = str(result_cell.value).lower() if result_cell.value else ""
                
                if value > threshold or (result_value and "fail" in result_value):
                    cell.fill = fail_fill
                    cell.font = fail_font
                    
                    harmonic_cell = ws.cell(row=r_idx, column=1)
                    harmonic_cell.fill = harmonic_fill
            except:
                continue

def create_excel_download(tables_data, filename):
    """Create Excel file with all tables with highlighting and split by odd/even harmonics"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for table_name, table_data in tables_data.items():
            if table_data:
                df = process_table_data(table_data, table_name)
                if not df.empty:
                    split_dfs = split_table(df)
                    
                    table_prefix = "I" if "Current" in table_name else "V"
                    table_suffix = "D" if "Daily" in table_name else "F"
                    
                    for limit in ["95", "99"]:
                        odd_df, even_df = split_dfs[limit]
                        
                        for df_data, suffix in [(odd_df, 'O'), (even_df, 'E')]:
                            if not df_data.empty:
                                sheet_name = f"H_{table_prefix}{table_suffix}_{limit}_{suffix}"[:31]
                                df_data.to_excel(writer, sheet_name=sheet_name, index=False)
                                
                                workbook = writer.book
                                worksheet = workbook[sheet_name]
                                highlight_fails_in_excel(df_data, worksheet, start_row=2)
    
    output.seek(0)
    return output.getvalue()

def parse_filename_for_sheet_name(filename):
    """Parse filename to extract day number and time period for concise sheet naming"""
    filename_upper = filename.upper()
    
    if "7" in filename_upper and "DAY" in filename_upper:
        return "7Days"
    
    day_pattern = re.search(r'DAY\s*(\d+)\s*(DAY|NIGHT)', filename_upper)
    if day_pattern:
        day_num = day_pattern.group(1)
        period = "D" if "DAY" in day_pattern.group(2) else "N"
        return f"{day_num}{period}"
    
    day_only_pattern = re.search(r'DAY\s*(\d+)', filename_upper)
    if day_only_pattern:
        return f"{day_only_pattern.group(1)}D"
    
    clean_name = re.sub(r'[^\w]', '', filename.replace('.pdf', ''))
    return clean_name[:4]

def get_table_abbreviation(table_name):
    """Convert full table names to abbreviations for concise sheet naming"""
    table_upper = table_name.upper()
    
    if "CURRENT" in table_upper:
        prefix = "I"
    elif "VOLTAGE" in table_upper:
        prefix = "V"
    else:
        prefix = "X"
    
    if "FULL TIME RANGE" in table_upper:
        suffix = "F"
    elif "DAILY" in table_upper:
        suffix = "D"
    else:
        suffix = "X"
    
    return f"{prefix}{suffix}"

def create_bulk_excel_download(all_files_data):
    """Create Excel file with all PDFs using concise sheet naming and highlighting"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_file_map = {}
        
        for file_name, tables_data in all_files_data.items():
            file_prefix = parse_filename_for_sheet_name(file_name)
            
            for table_name, table_data in tables_data.items():
                if table_data:
                    df = process_table_data(table_data, table_name)
                    if not df.empty:
                        table_abbrev = get_table_abbreviation(table_name)
                        sheet_name = f"{file_prefix}_H_{table_abbrev}"
                        
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        
                        original_sheet_name = sheet_name
                        counter = 1
                        while sheet_name in sheet_file_map:
                            sheet_name = f"{original_sheet_name}_{counter}"
                            if len(sheet_name) > 31:
                                truncated = original_sheet_name[:31-len(f"_{counter}")]
                                sheet_name = f"{truncated}_{counter}"
                            counter += 1
                        
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheet_file_map[sheet_name] = file_name
                        
                        workbook = writer.book
                        worksheet = workbook[sheet_name]
                        highlight_fails_in_excel(df, worksheet, start_row=2)
    
    output.seek(0)
    wb = load_workbook(output)
    
    for sheet_name, file_name in sheet_file_map.items():
        ws = wb[sheet_name]
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=f"File: {file_name}")
    
    output2 = BytesIO()
    wb.save(output2)
    output2.seek(0)
    return output2.getvalue()

def analyze_failures(df):
    """Identify and summarize all harmonic violations"""
    if df.empty:
        return pd.DataFrame()
    
    violations = []
    measured_cols = [col for col in df.columns if col.startswith('Measured_')]
    
    for _, row in df.iterrows():
        try:
            threshold = float(row['Reg Max[%]'])
            harmonic = int(row['N'])
            time_limit = row['[%]']
            
            for col in measured_cols:
                phase = col.split('_')[-1]  # Gets V1N/V2N/V3N or I1/I2/I3
                value = float(row[col])
                
                if value > threshold:
                    violations.append({
                        'Harmonic': harmonic,
                        'Phase': phase,
                        'Time Limit (%)': time_limit,
                        'Allowed (%)': threshold,
                        'Measured (%)': value,
                        'Exceedance (%)':value - threshold
                    })
        except (ValueError, KeyError):
            continue
    
    return pd.DataFrame(violations)

def display_metadata(component_text, block, feeder, company, report_info):
    """Display metadata in organized columns"""
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"- **Block:** `{block}`")
        st.markdown(f"- **Bay/Feeder:** `{feeder}`")
        st.markdown(f"- **Company:** `{company}`")
    with col2:
        st.markdown(f"- **Start Time:** `{report_info['start_time']}`")
        st.markdown(f"- **End Time:** `{report_info['end_time']}`")
    with col3:
        st.markdown(f"- **GMT Offset:** `{report_info['gmt']}`")
        st.markdown(f"- **Report Version:** `{report_info['version']}`")


def display_violation_summary(tables, filename):
    """Display harmonic violation summary section"""
    st.markdown("---")
    st.subheader("🚨 Harmonic Violation Summary")
    
    all_violations = []
    for table_name, table_data in tables.items():
        if table_data:
            df = process_table_data(table_data, table_name)
            if not df.empty:
                violations = analyze_failures(df)
                if not violations.empty:
                    violations['Table'] = table_name
                    all_violations.append(violations)
    
    if all_violations:
        combined_violations = pd.concat(all_violations)
        
        # Display statistics
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Total Violations", len(combined_violations))
  
        # Show detailed table
        st.dataframe(
            combined_violations.sort_values(['Exceedance (%)', 'Harmonic'], ascending=False),
            column_config={
                "Harmonic": "Harmonic",
                "Phase": "Phase",
                "Time Limit (%)": st.column_config.NumberColumn("Time Limit %"),
                "Allowed (%)": st.column_config.NumberColumn("Limit %"),
                "Measured (%)": st.column_config.NumberColumn("Measured %"),
                "Exceedance (%)": st.column_config.NumberColumn("Over Limit %"),
                "Table": "Source Table"
            },
            hide_index=True,
            use_container_width=True
        )
        
        # Add download button
        csv = combined_violations.to_csv(index=False)
        st.download_button(
            label="📥 Download Violation Report",
            data=csv,
            file_name=f"{filename.replace('.pdf', '')}_violations.csv",
            mime="text/csv"
        )
    else:
        st.success("✅ No harmonic violations detected")

def display_instructions():
    """Display usage instructions when no files are selected"""
    st.info(
        """
        ## 📋 Instructions:

        ### 1. **Upload Files:**
        - Upload all 15 files from your weekly report folder

        ### 2. **Week Report Structure (15 files):**
        - A **7 Days Summary Report**
        - 14 **Daily Reports** (Day 1-7, Day/Night)

        ### 3. **Tables Extracted:**
        - **Harmonic Voltage Full Time Range** (Pages 2-3)
        - **Harmonic Current Full Time Range** (Pages 8-10)
        - **Harmonic Voltage Daily** (Pages 5-6)
        - **Harmonic Current Daily** (Pages 12-13)

        ### 4. **Features:**
        - 📂 Click any **PDF file** from the sidebar to view its tables
        - 📊 Tables are split by **99% / 95% time limits** and **odd/even harmonics**
        - 🚨 **Violation Summary** shows all harmonic limit exceedances
        - 📥 Download individual PDF tables as **Excel**
        - 📦 **Bulk Download** processes all files at once

        ### 5. **Excel Download Format:**
        - Each table type appears on **separate sheets** for odd/even harmonics
        - Sheet naming: H_[V/I][F/D]_[95/99]_[O/E]
        - Bulk download includes all files with file names in sheet headers
        - Failed measurements are highlighted in red
        """
    )

# Main Processing Logic
if selected_file:
    filename = selected_file.name
    st.markdown(f"---\n### 📄 Processing: `{filename}`")
    
    try:
        # Extract and display metadata
        component_text, block, feeder, company, report_info = extract_metadata(selected_file, filename)
        display_metadata(component_text, block, feeder, company, report_info)
            
        # Extract and process tables
        tables = extract_tables_from_pdf(selected_file)
        
        # Download button for current file
        if any(tables.values()):
            excel_data = create_excel_download(tables, filename)
            st.download_button(
                label="📥 Download Current File Tables (Excel)",
                data=excel_data,
                file_name=f"{filename.replace('.pdf', '')}_tables.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Display violation summary
        display_violation_summary(tables, filename)
        
        # Display tables
        st.markdown("---")
        st.subheader("Harmonic Tables Overview")
        # Display tables organized by type
        voltage_tables = [name for name in SUPPORTED_TABLES if "Voltage" in name]
        current_tables = [name for name in SUPPORTED_TABLES if "Current" in name]

        # Voltage Section
        with st.expander("🔵 **VOLTAGE MEASUREMENTS**", expanded=True):
            for table_name in voltage_tables:
                if tables.get(table_name):
                    st.markdown(f"<h3 style='color: #4A90E2;'> {table_name}</h3>", unsafe_allow_html=True)
                    try:
                        df = process_table_data(tables[table_name], table_name)
                        if not df.empty:
                            split_dfs = split_table(df)
                            
                            for limit in ["95", "99"]:
                                odd_df, even_df = split_dfs[limit]
                                if not odd_df.empty or not even_df.empty:
                                    st.markdown(f"<h3 style='color: #6BAED6;'> {limit}% Time Limit</h3>", unsafe_allow_html=True)
                                    display_table_section("", odd_df, even_df)  
                        else:
                            st.warning(f"No valid data found in {table_name}")
                    except Exception as e:
                        st.error(f"Error processing table {table_name}: {str(e)}")
                else:
                    st.info(f"No data found for: {table_name}")

        st.markdown("---")

        # Current Section  
        with st.expander("🟠 **CURRENT MEASUREMENTS**", expanded=True):
            for table_name in current_tables:
                if tables.get(table_name):
                    st.markdown(f"<h3 style='color: #E78400;'> {table_name}</h3>", unsafe_allow_html=True)
                    try:
                        df = process_table_data(tables[table_name], table_name)
                        if not df.empty:
                            split_dfs = split_table(df)
                            
                            for limit in ["95", "99"]:
                                odd_df, even_df = split_dfs[limit]
                                if not odd_df.empty or not even_df.empty:
                                    st.markdown(f"<h3 style='color: #EABD8C;'> {limit}% Time Limit</h3>", unsafe_allow_html=True)
                                    display_table_section("", odd_df, even_df)  
                        else:
                            st.warning(f"No valid data found in {table_name}")
                    except Exception as e:
                        st.error(f"Error processing table {table_name}: {str(e)}")
                else:
                    st.info(f"No data found for: {table_name}")
            
    except Exception as e:
        st.error(f"Error processing file {filename}: {str(e)}")

# Bulk Download Section
if uploaded_files:
    st.markdown("---")
    st.subheader("📦 Bulk Download")
    
    if st.button("📥 Process All Files & Download Excel"):
        all_files_data = {}
        progress_bar = st.progress(0)
        
        for i, file in enumerate(uploaded_files):
            filename = file.name
            try:
                tables = extract_tables_from_pdf(file)
                if any(tables.values()):
                    all_files_data[filename] = tables
            except Exception as e:
                st.error(f"Error processing {filename}: {str(e)}")
            
            progress_bar.progress((i + 1) / len(uploaded_files))
        
        if all_files_data:
            bulk_excel_data = create_bulk_excel_download(all_files_data)
            st.download_button(
                label="📥 Download All Files Excel",
                data=bulk_excel_data,
                file_name="bulk_harmonic_reports.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"Processed {len(all_files_data)} files successfully!")
        else:
            st.warning("No valid data found in any files.")

# Display instructions when no files are uploaded
if not selected_file and not uploaded_files:
    display_instructions()